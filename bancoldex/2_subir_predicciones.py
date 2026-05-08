"""
PASO 2: Lee el Excel relleno y sube las predicciones a Supabase para el grupo Bancoldex.
Ejecutar: python3 2_subir_predicciones.py
Requiere: pip install requests openpyxl
"""

import requests
import openpyxl

SUPABASE_URL  = 'https://sjixxpdoeqxcslcuyemf.supabase.co'
SUPABASE_KEY  = 'sb_publishable_tSYbG9MpWR1QBDmM1zwr4Q_A7rhQK5S'
GRUPO         = 'Bancoldex'
ARCHIVO_EXCEL = 'predicciones_bancoldex.xlsx'

HEADERS = {
    'apikey': SUPABASE_KEY,
    'Authorization': f'Bearer {SUPABASE_KEY}',
    'Content-Type': 'application/json',
}

def get(tabla, params=''):
    r = requests.get(f'{SUPABASE_URL}/rest/v1/{tabla}?{params}', headers=HEADERS)
    r.raise_for_status()
    return r.json()

def post(tabla, body):
    r = requests.post(f'{SUPABASE_URL}/rest/v1/{tabla}', json=body, headers=HEADERS)
    r.raise_for_status()
    return r.json()

def upsert(tabla, body, on_conflict):
    h = {**HEADERS, 'Prefer': 'resolution=merge-duplicates,return=representation'}
    r = requests.post(f'{SUPABASE_URL}/rest/v1/{tabla}?on_conflict={on_conflict}', json=body, headers=h)
    r.raise_for_status()
    return r.json()

def obtener_o_crear_participante(nombre):
    nombre = nombre.strip()
    data = get('participantes', f"nombre=ilike.{requests.utils.quote(nombre)}&grupo=eq.{requests.utils.quote(GRUPO)}&select=id,nombre")
    if data:
        print(f'   Ya existe en la base de datos')
        return data[0]
    resultado = post('participantes', {'nombre': nombre, 'grupo': GRUPO})
    if isinstance(resultado, list):
        return resultado[0]
    return resultado

def main():
    print(f'Leyendo {ARCHIVO_EXCEL}...')
    wb = openpyxl.load_workbook(ARCHIVO_EXCEL, data_only=True)
    ws = wb.active

    # Fila 2: nombres de participantes (columnas H en adelante, cada 3 columnas)
    # Fila 4 (oculta): partido_ids por columna
    # Filas 5+: predicciones (col_l = local, col_v = visitante)

    COL_INICIO = 8  # columna H
    FILA_IDS   = 4
    FILA_NOMBRES = 2
    FILA_DATOS = 5

    # Detectar participantes (cada 3 columnas desde H)
    participantes_cols = []
    col = COL_INICIO
    while col <= ws.max_column:
        nombre = ws.cell(FILA_NOMBRES, col).value
        if nombre and str(nombre).strip() not in ('', 'Local', 'Visit.', 'Pts'):
            participantes_cols.append((str(nombre).strip(), col))
        col += 3

    if not participantes_cols:
        print('❌ No se encontraron participantes. ¿Están los nombres en la fila 2?')
        return

    print(f'  Participantes encontrados: {[n for n,_ in participantes_cols]}')

    # Detectar partidos: leer fila 4 (IDs ocultos) junto con los datos
    # Para cada fila de datos, la col_l de cada participante tiene el mismo partido_id en fila 4
    total_preds = 0
    total_personas = 0

    for nombre, col_l in participantes_cols:
        col_v   = col_l + 1

        print(f'\n👤 {nombre}')
        participante = obtener_o_crear_participante(nombre)
        part_id = participante['id']
        total_personas += 1

        predicciones = []
        fila = FILA_DATOS
        while fila <= ws.max_row:
            # Leer partido_id desde fila 4 de esa columna
            partido_id = ws.cell(FILA_IDS, col_l).value

            # Los IDs están almacenados en fila 4, misma columna que local pred
            # pero el ID es el mismo para todas las filas de ese partido
            # En realidad los IDs están repetidos por partido en fila 4...
            # Obtener el ID específico de este partido desde columna A de fila 4
            # Mejor: leer el ID desde la columna A de fila 4 para este fila de partido

            # El partido_id está en fila 4 para la columna del primer participante
            # Necesitamos leerlo de la fila correcta
            # En el script de generación, pusimos ws.cell(4, col_l, p['id']) por cada partido en el loop de filas
            # pero eso sobreescribe el mismo cell(4, col_l) para cada partido — error de diseño
            # En su lugar, vamos a leer el partido desde columna A (número de partido) y buscarlo por número

            val_local  = ws.cell(fila, col_l).value
            val_visit  = ws.cell(fila, col_v).value

            # Detectar fin de datos (separador o vacío total)
            val_a = ws.cell(fila, 1).value
            if val_a is None or str(val_a).startswith('—'):
                break

            try:
                gl = int(val_local) if val_local is not None and str(val_local).strip() != '' else None
                gv = int(val_visit) if val_visit is not None and str(val_visit).strip() != '' else None
            except (ValueError, TypeError):
                gl, gv = None, None

            # Obtener partido_id desde Supabase usando el número de partido (col A)
            num_partido = ws.cell(fila, 1).value
            if gl is not None and gv is not None and num_partido:
                predicciones.append({
                    'numero': int(num_partido),
                    'gl': gl,
                    'gv': gv,
                })

            fila += 1

        if not predicciones:
            print(f'   ⚠️  Sin predicciones')
            continue

        # Obtener IDs reales de partidos por número
        numeros = [p['numero'] for p in predicciones]
        partidos_db = get('partidos', f"numero=in.({','.join(map(str,numeros))})&select=id,numero")
        id_por_numero = {p['numero']: p['id'] for p in partidos_db}

        rows = []
        for pred in predicciones:
            pid = id_por_numero.get(pred['numero'])
            if pid:
                rows.append({
                    'participante_id': part_id,
                    'partido_id': pid,
                    'goles_local': pred['gl'],
                    'goles_visitante': pred['gv'],
                })

        if rows:
            upsert('predicciones', rows, 'participante_id,partido_id')
            print(f'   ✅ {len(rows)} predicciones subidas')
            total_preds += len(rows)
        else:
            print(f'   ⚠️  Sin predicciones válidas')

    print(f'\n🏁 Listo: {total_personas} personas, {total_preds} predicciones subidas al grupo {GRUPO}.')

if __name__ == '__main__':
    main()
