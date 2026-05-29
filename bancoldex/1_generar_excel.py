"""
Genera el Excel de predicciones por persona.
- Formato: Equipo Local [tu pred] vs [tu pred] Equipo Visitante
- Validación numérica en celdas de predicción (solo enteros 0-99)
- Hoja protegida: solo se pueden editar predicciones, penaltis y nombre
- Formato condicional: verde/azul/rojo según clasificación
- Tabla de mejores terceros (top 8 de 12 grupos)
- Fases eliminatorias con celda de penaltis para empates

Ejecutar:  python3 1_generar_excel.py
Requiere:  pip install requests openpyxl
"""

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from collections import defaultdict, OrderedDict
from datetime import datetime

SUPABASE_URL  = 'https://sjixxpdoeqxcslcuyemf.supabase.co'
SUPABASE_KEY  = 'sb_publishable_tSYbG9MpWR1QBDmM1zwr4Q_A7rhQK5S'
ARCHIVO       = 'predicciones_bancoldex.xlsx'

# Bracket oficial FIFA 2026 (fuente: schedule oficial FIFA, M73-M88)
R32_BRACKET = [
    (('G',2,'A'), ('G',2,'B')),   # M73
    (('G',1,'E'), ('T3',1)),       # M74
    (('G',1,'F'), ('G',2,'C')),   # M75
    (('G',1,'C'), ('G',2,'F')),   # M76
    (('G',1,'I'), ('T3',2)),       # M77
    (('G',2,'E'), ('G',2,'I')),   # M78
    (('G',1,'A'), ('T3',3)),       # M79
    (('G',1,'L'), ('T3',4)),       # M80
    (('G',1,'D'), ('T3',5)),       # M81
    (('G',1,'G'), ('T3',6)),       # M82
    (('G',2,'K'), ('G',2,'L')),   # M83
    (('G',1,'H'), ('G',2,'J')),   # M84
    (('G',1,'B'), ('T3',7)),       # M85
    (('G',1,'J'), ('G',2,'H')),   # M86
    (('G',1,'K'), ('T3',8)),       # M87
    (('G',2,'D'), ('G',2,'G')),   # M88
]
# R16: M89=W74vsW77, M90=W73vsW75, M91=W83vsW84, M92=W81vsW82
#      M93=W76vsW78, M94=W79vsW80, M95=W86vsW88, M96=W85vsW87
# Índices en r32_filas: M73=0,M74=1,M75=2,M76=3,M77=4,M78=5,M79=6,M80=7,M81=8,M82=9,M83=10,M84=11,M85=12,M86=13,M87=14,M88=15
R16_PAIRS = [(1,4),(0,2),(10,11),(8,9),(3,5),(6,7),(13,15),(12,14)]
QF_PAIRS  = [(0,1),(2,3),(4,5),(6,7)]   # M97=W89vsW90, M98=W91vsW92, M99=W93vsW94, M100=W95vsW96
SF_PAIRS  = [(0,1),(2,3)]               # M101=W97vsW98, M102=W99vsW100

# ── Columnas ──────────────────────────────────────────────────────────────────
C_NUM   = 1; C_FECHA = 2; C_HORA  = 3; C_LOCAL = 4
C_PL    = 5   # E  predicción local  (amarillo, editable)
C_VS    = 6; C_PV    = 7   # G  predicción visitante (amarillo, editable)
C_VISIT = 8; C_REAL  = 9
C_PEN   = 10  # J  ganador penaltis  (lila, editable solo en eliminatorias)
C_ST_EQ  = 11; C_ST_PTS = 12; C_ST_GF  = 13
C_ST_GC  = 14; C_ST_DG  = 15; C_ST_SC  = 16; C_ST_RK  = 17
C_FP     = 18  # R — desempate fairplay (editable, 1-4)

# ── Estilos ───────────────────────────────────────────────────────────────────
def fill(c): return PatternFill('solid', fgColor=c)
def borde(c='cccccc'):
    s = Side(style='thin', color=c)
    return Border(left=s, right=s, top=s, bottom=s)
def font(bold=False, size=10, color='111111', italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic, name='Calibri')
def alin(h='center', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

VERDE  = '1a5c2a'; VERDE2 = '2d7a3f'; VERDE3 = '3d8b50'
DORADO = 'c9a84c'; ORO_L  = 'fffde7'
GRIS1  = 'f5f5f5'; GRIS2  = 'e8e8e8'
BLANCO = 'FFFFFF'; NEGRO  = '000000'
AZUL_T3 = '0d47a1'; AZUL_T3_L = 'e8f0fe'
LILA_L  = 'ede7f6'

RONDA_COLOR = {
    '16avos':     ('1565C0', 'e3f0ff'),
    'Octavos':    ('6A1F8A', 'f3e5f5'),
    'Cuartos':    ('E65100', 'fff3e0'),
    'Semis':      ('880E4F', 'fce4ec'),
    '3er Puesto': ('4E342E', 'efebe9'),
    'Final':      (DORADO,   ORO_L),
}

# ── Helpers ───────────────────────────────────────────────────────────────────
def col(n): return get_column_letter(n)
def ref(fila, c): return f'{col(c)}{fila}'

def f_ganador(fila):
    pl = ref(fila, C_PL); pv = ref(fila, C_PV)
    cl = ref(fila, C_LOCAL); cv = ref(fila, C_VISIT)
    pen = ref(fila, C_PEN)
    return (f'=IFERROR(IF(AND(ISNUMBER({pl}),ISNUMBER({pv})),'
            f'IF({pl}>{pv},{cl},'
            f'IF({pv}>{pl},{cv},'
            f'IF({pen}<>"",{pen},"→ Pon ganador penaltis en col J"))),'
            f'"Por definir"),"")')

def f_perdedor(fila):
    pl = ref(fila, C_PL); pv = ref(fila, C_PV)
    cl = ref(fila, C_LOCAL); cv = ref(fila, C_VISIT)
    pen = ref(fila, C_PEN)
    return (f'=IFERROR(IF(AND(ISNUMBER({pl}),ISNUMBER({pv})),'
            f'IF({pl}>{pv},{cv},'
            f'IF({pv}>{pl},{cl},'
            f'IF({pen}={cl},{cv},IF({pen}={cv},{cl},"Por definir")))),'
            f'"Por definir"),"")')

# ── Supabase ──────────────────────────────────────────────────────────────────
def fetch(tabla, params=''):
    r = requests.get(f'{SUPABASE_URL}/rest/v1/{tabla}?{params}',
        headers={'apikey': SUPABASE_KEY, 'Authorization': f'Bearer {SUPABASE_KEY}'})
    r.raise_for_status(); return r.json()

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print('Obteniendo partidos de Supabase...')
    todos = fetch('partidos',
        'select=id,numero,fase,grupo,equipo_local,equipo_visitante,fecha,hora,orden&order=orden.asc')
    grupos_fase = [p for p in todos if p['fase'] == 'grupos']
    print(f'  {len(grupos_fase)} partidos de grupos')

    por_grupo = defaultdict(list)
    for p in grupos_fase:
        por_grupo[p['grupo']].append(p)
    letras = sorted(por_grupo.keys())

    wb = openpyxl.Workbook()
    wb.calculation.calcMode = 'auto'
    ws = wb.active
    ws.title = 'Predicciones'
    ws.freeze_panes = 'E3'

    # Listas para rastrear celdas editables
    pred_refs = []   # predicciones (validación numérica + desbloquear)
    pen_refs  = []   # penaltis (solo desbloquear, texto libre)
    fp_refs   = []   # desempate fairplay (validación 1-4 + desbloquear)

    # ── Anchos ───────────────────────────────────────────────────────────────
    for c_, w in {1:5, 2:14, 3:8, 4:26, 5:6, 6:4, 7:6, 8:26, 9:12,
                  10:20, 11:26, 12:5, 13:5, 14:5, 15:5, 16:5, 18:22}.items():
        ws.column_dimensions[col(c_)].width = w

    ws.column_dimensions[col(C_ST_SC)].hidden = True
    ws.column_dimensions[col(C_ST_RK)].hidden = True

    # ── Fila 1: título + nombre ───────────────────────────────────────────────
    ws.merge_cells('A1:I1')
    c = ws['A1']
    c.value = '⚽  POLLA MUNDIAL 2026  –  BANCOLDEX'
    c.font = font(bold=True, size=14, color=BLANCO)
    c.fill = fill(VERDE); c.alignment = alin()
    ws.row_dimensions[1].height = 28

    ws['J1'].value = ''; ws['J1'].fill = fill(VERDE)

    ws.merge_cells('K1:N1')
    ws['K1'].value = 'Tu nombre:'; ws['K1'].font = font(bold=True, size=10, color=BLANCO)
    ws['K1'].fill = fill(VERDE); ws['K1'].alignment = alin('right')
    ws.column_dimensions['O'].width = 22
    c = ws['O1']
    c.value = '👉 Escribe tu nombre aquí'
    c.font = font(bold=True, size=10, color='555555', italic=True)
    c.fill = fill(ORO_L); c.alignment = alin('left'); c.border = borde(DORADO)

    # ── Fila 2: cabeceras ─────────────────────────────────────────────────────
    for c_, txt, bg, fg in [
        (C_NUM,  'Nº',           GRIS2, '555555'),
        (C_FECHA,'Fecha',        GRIS2, '555555'),
        (C_HORA, 'Hora',         GRIS2, '555555'),
        (C_LOCAL,'Equipo local', VERDE2, BLANCO),
        (C_PL,   '← tu pred',   DORADO, NEGRO),
        (C_VS,   'vs',           GRIS1, '888888'),
        (C_PV,   'tu pred →',   DORADO, NEGRO),
        (C_VISIT,'Equipo visit.',VERDE2, BLANCO),
        (C_REAL, 'Resultado real',GRIS2,'555555'),
        (C_PEN,  'Penaltis (si empate)', '7b1fa2', BLANCO),
    ]:
        c = ws.cell(2, c_, txt)
        c.font = font(bold=True, size=9, color=fg)
        c.fill = fill(bg); c.alignment = alin(wrap=True); c.border = borde()
    ws.row_dimensions[2].height = 46

    for c_, txt in [(C_ST_EQ,'Clasificación'), (C_ST_PTS,'Pts'),
                    (C_ST_GF,'GF'), (C_ST_GC,'GC'), (C_ST_DG,'DG')]:
        c = ws.cell(2, c_, txt)
        c.font = font(bold=True, size=8, color='555555')
        c.fill = fill(GRIS2); c.alignment = alin(); c.border = borde()
    c = ws.cell(2, C_FP, '⚖️ DESEMPATE FAIRPLAY\nSolo si empate total en pts+DG+GF\nEscribe: 1°→1  2°→2  3°→3  4°→4')
    c.font = font(bold=True, size=8, color='4a148c')
    c.fill = fill('ede7f6'); c.alignment = alin(wrap=True); c.border = borde('ce93d8')

    # ── Escribir grupos ───────────────────────────────────────────────────────
    fila = 3
    match_filas = {}
    grupo_st    = {}

    for letra in letras:
        matches = por_grupo[letra]

        ws.merge_cells(f'A{fila}:J{fila}')
        c = ws.cell(fila, 1, f'  GRUPO {letra}')
        c.font = font(bold=True, size=10, color=BLANCO)
        c.fill = fill(VERDE3); c.alignment = alin('left')
        ws.merge_cells(f'K{fila}:P{fila}')
        c = ws.cell(fila, C_ST_EQ, f'Clasificación Grupo {letra}')
        c.font = font(bold=True, size=9, color=BLANCO)
        c.fill = fill(VERDE3); c.alignment = alin()
        ws.row_dimensions[fila].height = 20
        fila += 1

        for idx, m in enumerate(matches):
            match_filas[m['id']] = fila
            par = idx % 2 == 0
            bg  = 'f9fafb' if par else BLANCO
            try:
                fd = datetime.strptime(m['fecha'], '%Y-%m-%d').strftime('%a %d %b') if m.get('fecha') else ''
            except: fd = m.get('fecha','')
            hora = (m.get('hora') or '')[:5]

            for c_, val, al in [
                (C_NUM,   m.get('numero',''), alin()),
                (C_FECHA, fd,                 alin('left')),
                (C_HORA,  hora,               alin()),
                (C_LOCAL, m['equipo_local'],   alin('right')),
                (C_PL,    '',                  alin()),
                (C_VS,    'vs',                alin()),
                (C_PV,    '',                  alin()),
                (C_VISIT, m['equipo_visitante'],alin('left')),
                (C_REAL,  '',                  alin()),
                (C_PEN,   '',                  alin()),
            ]:
                c = ws.cell(fila, c_, val)
                c.alignment = al; c.border = borde()
                if c_ in (C_PL, C_PV):
                    c.font = font(bold=True, size=13); c.fill = fill(ORO_L)
                elif c_ == C_REAL:
                    c.font = font(size=9, color='888888'); c.fill = fill(GRIS1)
                elif c_ == C_PEN:
                    c.fill = fill(GRIS1)
                else:
                    c.font = font(size=9); c.fill = fill(bg)

            # Registrar celdas editables de este partido
            pred_refs.append(ref(fila, C_PL))
            pred_refs.append(ref(fila, C_PV))
            ws.row_dimensions[fila].height = 18
            fila += 1

        # ── Tabla de clasificación ────────────────────────────────────────────
        equipos = list(OrderedDict.fromkeys(
            [m['equipo_local'] for m in matches] +
            [m['equipo_visitante'] for m in matches]
        ))[:4]
        mf = {m['id']: match_filas[m['id']] for m in matches}
        grupo_st[letra] = fila

        for eq_idx, eq in enumerate(equipos):
            st = fila + eq_idx
            c = ws.cell(st, C_ST_EQ, eq)
            c.font = font(size=9, bold=True); c.fill = fill(GRIS1)
            c.alignment = alin('left'); c.border = borde()

            cl = [m for m in matches if m['equipo_local']     == eq]
            cv = [m for m in matches if m['equipo_visitante'] == eq]

            partes_pts = []
            for m in cl:
                r = mf[m['id']]
                partes_pts.append(f'IF(AND(ISNUMBER({ref(r,C_PL)}),ISNUMBER({ref(r,C_PV)})),IF({ref(r,C_PL)}>{ref(r,C_PV)},3,IF({ref(r,C_PL)}={ref(r,C_PV)},1,0)),0)')
            for m in cv:
                r = mf[m['id']]
                partes_pts.append(f'IF(AND(ISNUMBER({ref(r,C_PL)}),ISNUMBER({ref(r,C_PV)})),IF({ref(r,C_PV)}>{ref(r,C_PL)},3,IF({ref(r,C_PL)}={ref(r,C_PV)},1,0)),0)')

            partes_gf = []
            for m in cl:
                r = mf[m['id']]
                partes_gf.append(f'IF(ISNUMBER({ref(r,C_PL)}),{ref(r,C_PL)},0)')
            for m in cv:
                r = mf[m['id']]
                partes_gf.append(f'IF(ISNUMBER({ref(r,C_PV)}),{ref(r,C_PV)},0)')

            partes_gc = []
            for m in cl:
                r = mf[m['id']]
                partes_gc.append(f'IF(ISNUMBER({ref(r,C_PV)}),{ref(r,C_PV)},0)')
            for m in cv:
                r = mf[m['id']]
                partes_gc.append(f'IF(ISNUMBER({ref(r,C_PL)}),{ref(r,C_PL)},0)')

            pts_f = ref(st, C_ST_PTS); gf_f = ref(st, C_ST_GF)
            gc_f  = ref(st, C_ST_GC);  dg_f = ref(st, C_ST_DG)
            fp_f  = ref(st, C_FP)

            ws.cell(st, C_ST_PTS).value = '=' + '+'.join(partes_pts)
            ws.cell(st, C_ST_GF ).value = '=' + '+'.join(partes_gf)
            ws.cell(st, C_ST_GC ).value = '=' + '+'.join(partes_gc)
            ws.cell(st, C_ST_DG ).value = f'={gf_f}-{gc_f}'

            # Si el usuario llenó el desempate fairplay, usarlo como tiebreaker final
            score = (f'IF(ISNUMBER({fp_f}),'
                     f'{pts_f}*10000+{dg_f}*100+{gf_f}+(1-{fp_f}/10000),'
                     f'{pts_f}*10000+{dg_f}*100+{gf_f}+{(eq_idx+1)*0.001:.3f})')
            ws.cell(st, C_ST_SC).value = f'={score}'
            ws.cell(st, C_ST_SC).font  = font(size=8, color='aaaaaa')

            for cx in [C_ST_PTS, C_ST_GF, C_ST_GC, C_ST_DG]:
                ws.cell(st, cx).fill      = fill(GRIS1)
                ws.cell(st, cx).alignment = alin()
                ws.cell(st, cx).border    = borde()

            # Celda de desempate fairplay
            fp_cell = ws.cell(st, C_FP)
            fp_cell.fill = fill('ede7f6'); fp_cell.alignment = alin()
            fp_cell.border = borde('ce93d8')
            fp_cell.font = font(size=9, color='4a148c', bold=True)
            fp_cell.protection = Protection(locked=False)
            fp_refs.append(fp_f)

            ws.row_dimensions[st].height = 17

        score_range_full = f'{col(C_ST_SC)}{fila}:{col(C_ST_SC)}{fila+3}'
        for eq_idx in range(4):
            st = fila + eq_idx
            sc = ref(st, C_ST_SC)
            ws.cell(st, C_ST_RK).value = f'=RANK({sc},{score_range_full},0)'
            ws.cell(st, C_ST_RK).font  = font(size=8, color='aaaaaa')

        # ── Formato condicional por rank ──────────────────────────────────────
        rng_cf  = f'K{fila}:O{fila+3}'
        q_letra = col(C_ST_RK)
        l_letra = col(C_ST_PTS)
        has_preds = f'SUM(${l_letra}${fila}:${l_letra}${fila+3})>0'

        ws.conditional_formatting.add(rng_cf, FormulaRule(
            formula=[f'AND(${q_letra}{fila}<=2,{has_preds})'],
            fill=PatternFill('solid', fgColor='c8e6c9'),
            font=Font(bold=True, color='1b5e20', name='Calibri')
        ))
        ws.conditional_formatting.add(rng_cf, FormulaRule(
            formula=[f'AND(${q_letra}{fila}=3,{has_preds})'],
            fill=PatternFill('solid', fgColor='bbdefb'),
            font=Font(bold=True, color='0d47a1', name='Calibri')
        ))
        ws.conditional_formatting.add(rng_cf, FormulaRule(
            formula=[f'AND(${q_letra}{fila}=4,{has_preds})'],
            fill=PatternFill('solid', fgColor='ffcdd2'),
            font=Font(color='c62828', name='Calibri')
        ))

        fila += 4 + 1

    # ── Función para 1° o 2° de un grupo ─────────────────────────────────────
    def f_pos_grupo(letra, pos):
        fi = grupo_st[letra]
        eq_rng   = f'{col(C_ST_EQ)}{fi}:{col(C_ST_EQ)}{fi+3}'
        rank_rng = f'{col(C_ST_RK)}{fi}:{col(C_ST_RK)}{fi+3}'
        pts_rng  = f'{col(C_ST_PTS)}{fi}:{col(C_ST_PTS)}{fi+3}'
        return (f'=IF(SUM({pts_rng})=0,"Por definir",'
                f'IFERROR(INDEX({eq_rng},MATCH({pos},{rank_rng},0)),"Por definir"))')

    # ── Tabla de Mejores Terceros ─────────────────────────────────────────────
    fila += 1

    ws.merge_cells(f'A{fila}:J{fila}')
    c = ws.cell(fila, 1, '  MEJORES TERCEROS — ranking de los 12 grupos (top 8 clasifican a 16avos)')
    c.font = font(bold=True, size=10, color=BLANCO)
    c.fill = fill(AZUL_T3); c.alignment = alin('left')
    ws.merge_cells(f'K{fila}:O{fila}')
    c = ws.cell(fila, C_ST_EQ, '🟢 Clasifica   🔵 Posible 3°   🔴 Eliminado')
    c.font = font(bold=True, size=9, color=BLANCO)
    c.fill = fill(AZUL_T3); c.alignment = alin()
    ws.row_dimensions[fila].height = 20
    fila += 1

    for c_, txt in [(C_ST_EQ,'Equipo (3°)'), (C_ST_PTS,'Pts'),
                    (C_ST_GF,'GF'), (C_ST_GC,'GC'), (C_ST_DG,'DG')]:
        c = ws.cell(fila, c_, txt)
        c.font = font(bold=True, size=8, color='555555')
        c.fill = fill(GRIS2); c.alignment = alin(); c.border = borde()
    ws.row_dimensions[fila].height = 17
    fila += 1

    terceros_fila_inicio = fila
    t3_sc_rng = f'{col(C_ST_SC)}{terceros_fila_inicio}:{col(C_ST_SC)}{terceros_fila_inicio+11}'
    t3_rk_rng = f'{col(C_ST_RK)}{terceros_fila_inicio}:{col(C_ST_RK)}{terceros_fila_inicio+11}'
    t3_eq_rng = f'{col(C_ST_EQ)}{terceros_fila_inicio}:{col(C_ST_EQ)}{terceros_fila_inicio+11}'

    for g_idx, letra in enumerate(letras):
        fi = grupo_st[letra]
        eq_rng   = f'{col(C_ST_EQ)}{fi}:{col(C_ST_EQ)}{fi+3}'
        rank_rng = f'{col(C_ST_RK)}{fi}:{col(C_ST_RK)}{fi+3}'
        pts_rng  = f'{col(C_ST_PTS)}{fi}:{col(C_ST_PTS)}{fi+3}'
        gf_rng   = f'{col(C_ST_GF)}{fi}:{col(C_ST_GF)}{fi+3}'
        gc_rng   = f'{col(C_ST_GC)}{fi}:{col(C_ST_GC)}{fi+3}'
        dg_rng   = f'{col(C_ST_DG)}{fi}:{col(C_ST_DG)}{fi+3}'

        t = terceros_fila_inicio + g_idx

        c = ws.cell(t, C_NUM, f'G{letra}')
        c.font = font(size=8, bold=True, color='555555')
        c.fill = fill(GRIS2); c.alignment = alin(); c.border = borde()

        eq_f = (f'=IF(SUM({pts_rng})=0,"Por definir",'
                f'IFERROR(INDEX({eq_rng},MATCH(3,{rank_rng},0)),"Por definir"))')
        c = ws.cell(t, C_ST_EQ, eq_f)
        c.font = font(size=9); c.fill = fill(GRIS1)
        c.alignment = alin('left'); c.border = borde()

        def t3_val(rng):
            return (f'=IF(SUM({pts_rng})=0,0,'
                    f'IFERROR(INDEX({rng},MATCH(3,{rank_rng},0)),0))')

        for cx, rng in [(C_ST_PTS, pts_rng), (C_ST_GF, gf_rng),
                        (C_ST_GC, gc_rng),   (C_ST_DG, dg_rng)]:
            c = ws.cell(t, cx, t3_val(rng))
            if cx == C_ST_DG:
                gf_c = ref(t, C_ST_GF); gc_c = ref(t, C_ST_GC)
                c.value = f'=IF(SUM({pts_rng})=0,0,{gf_c}-{gc_c})'
            c.font = font(size=9); c.fill = fill(GRIS1)
            c.alignment = alin(); c.border = borde()

        pts_c = ref(t, C_ST_PTS); dg_c = ref(t, C_ST_DG); gf_c = ref(t, C_ST_GF)
        tb = f'{(g_idx+1)*0.001:.3f}'
        ws.cell(t, C_ST_SC).value = (
            f'=IF(SUM({pts_rng})=0,-99999+{tb},'
            f'{pts_c}*10000+{dg_c}*100+{gf_c}+{tb})')
        ws.cell(t, C_ST_SC).font = font(size=8, color='aaaaaa')
        ws.row_dimensions[t].height = 17

    for g_idx in range(12):
        t = terceros_fila_inicio + g_idx
        sc_ref = ref(t, C_ST_SC)
        ws.cell(t, C_ST_RK).value = f'=RANK({sc_ref},{t3_sc_rng},0)'
        ws.cell(t, C_ST_RK).font  = font(size=8, color='aaaaaa')

    rng_t3  = f'K{terceros_fila_inicio}:O{terceros_fila_inicio+11}'
    q_letra = col(C_ST_RK)
    ws.conditional_formatting.add(rng_t3, FormulaRule(
        formula=[f'AND(${q_letra}{terceros_fila_inicio}<=8,${col(C_ST_SC)}{terceros_fila_inicio}>-99000)'],
        fill=PatternFill('solid', fgColor='bbdefb'),
        font=Font(bold=True, color='0d47a1', name='Calibri')
    ))

    fila = terceros_fila_inicio + 12

    def f_mejor_tercero(n):
        return (f'=IF(LARGE({t3_sc_rng},{n})<0,"Por definir",'
                f'IFERROR(INDEX({t3_eq_rng},MATCH({n},{t3_rk_rng},0)),"Por definir"))')

    # ── Fase eliminatoria ─────────────────────────────────────────────────────
    fila += 2

    def seccion_header(titulo, dark):
        nonlocal fila
        ws.merge_cells(f'A{fila}:P{fila}')
        c = ws.cell(fila, 1, f'  {titulo}')
        c.font = font(bold=True, size=11, color=BLANCO)
        c.fill = fill(dark); c.alignment = alin('left')
        ws.row_dimensions[fila].height = 24
        fila += 1

    def partido_elim(num, fl, fv, dark, light, nota=''):
        nonlocal fila
        for c_, val, al in [
            (C_NUM,   num,  alin()),
            (C_FECHA, nota, alin('left')),
            (C_HORA,  '',   alin()),
            (C_LOCAL, fl,   alin('right')),
            (C_PL,    '',   alin()),
            (C_VS,    'vs', alin()),
            (C_PV,    '',   alin()),
            (C_VISIT, fv,   alin('left')),
            (C_REAL,  '',   alin()),
            (C_PEN,   '',   alin()),
        ]:
            c = ws.cell(fila, c_)
            c.value = val
            c.alignment = al; c.border = borde()
            if c_ in (C_PL, C_PV):
                c.font = font(bold=True, size=13); c.fill = fill(ORO_L)
            elif c_ == C_PEN:
                c.font = font(size=9, color='4a148c', italic=True)
                c.fill = fill(LILA_L); c.border = borde('ce93d8')
            elif c_ == C_REAL:
                c.font = font(size=9, color='888888'); c.fill = fill(GRIS1)
            elif c_ in (C_LOCAL, C_VISIT):
                c.font = font(bold=True, size=9, italic=True, color='222222')
                c.fill = fill(light)
            else:
                c.font = font(size=9, color='555555'); c.fill = fill(light)

        # Registrar celdas editables de este partido eliminatorio
        pred_refs.append(ref(fila, C_PL))
        pred_refs.append(ref(fila, C_PV))
        pen_refs.append(ref(fila, C_PEN))

        ws.row_dimensions[fila].height = 18
        f = fila; fila += 1
        return f

    seccion_header('16AVOS DE FINAL', RONDA_COLOR['16avos'][0])
    r32_filas = []
    for i, (sl, sv) in enumerate(R32_BRACKET):
        fl = f_pos_grupo(sl[2], sl[1]) if sl[0]=='G' else f_mejor_tercero(sl[1])
        fv = f_pos_grupo(sv[2], sv[1]) if sv[0]=='G' else f_mejor_tercero(sv[1])
        r32_filas.append(partido_elim(i+1, fl, fv, *RONDA_COLOR['16avos'], nota=f'Partido {i+1}'))
    fila += 1

    seccion_header('OCTAVOS DE FINAL', RONDA_COLOR['Octavos'][0])
    r16_filas = []
    for i, (a, b) in enumerate(R16_PAIRS):
        r16_filas.append(partido_elim(i+1, f_ganador(r32_filas[a]), f_ganador(r32_filas[b]),
                                      *RONDA_COLOR['Octavos'], nota=f'Partido {i+1}'))
    fila += 1

    seccion_header('CUARTOS DE FINAL', RONDA_COLOR['Cuartos'][0])
    qf_filas = []
    for i, (a, b) in enumerate(QF_PAIRS):
        qf_filas.append(partido_elim(i+1, f_ganador(r16_filas[a]), f_ganador(r16_filas[b]),
                                     *RONDA_COLOR['Cuartos'], nota=f'Partido {i+1}'))
    fila += 1

    seccion_header('SEMIFINALES', RONDA_COLOR['Semis'][0])
    sf_filas = []
    for i, (a, b) in enumerate(SF_PAIRS):
        sf_filas.append(partido_elim(i+1, f_ganador(qf_filas[a]), f_ganador(qf_filas[b]),
                                     *RONDA_COLOR['Semis']))
    fila += 1

    seccion_header('3ER Y 4TO PUESTO', RONDA_COLOR['3er Puesto'][0])
    partido_elim(1, f_perdedor(sf_filas[0]), f_perdedor(sf_filas[1]), *RONDA_COLOR['3er Puesto'])
    fila += 1

    seccion_header('🏆 GRAN FINAL', RONDA_COLOR['Final'][0])
    partido_elim(1, f_ganador(sf_filas[0]), f_ganador(sf_filas[1]), *RONDA_COLOR['Final'])
    fila += 2

    # ── Leyenda ───────────────────────────────────────────────────────────────
    ws.merge_cells(f'A{fila}:J{fila}')
    c = ws.cell(fila, 1,
        '  🟢 Clasifica a 16avos   🔵 Posible mejor 3°   🔴 Eliminado   🟣 Escribe ganador si predices empate en eliminatoria   ⚖️ Col R: si hay empate total en el grupo, escribe 1 para el equipo que quieres primero, 2 para segundo…')
    c.font = font(size=8, color='555555', italic=True)
    c.fill = fill(GRIS1); c.alignment = alin('left')
    ws.row_dimensions[fila].height = 16
    fila += 2

    # ── Tabla de puntuación ───────────────────────────────────────────────────
    ws.merge_cells(f'A{fila}:I{fila}')
    c = ws.cell(fila, 1, 'SISTEMA DE PUNTUACIÓN')
    c.font = font(bold=True, size=10, color=BLANCO)
    c.fill = fill(VERDE); c.alignment = alin(); ws.row_dimensions[fila].height = 20
    fila += 1

    for i, (desc, pts) in enumerate([
        ('🎯 Marcador exacto',           '4 pts'),
        ('✓+ Resultado + un marcador',    '3 pts'),
        ('✓  Solo resultado correcto',     '2 pts'),
        ('~  Un marcador, resultado mal',  '1 pt'),
        ('✗  Ninguna',                     '0 pts'),
        ('',''),
        ('Bonus 16avos de final',          '+2 pts'),
        ('Bonus Octavos de final',         '+3 pts'),
        ('Bonus Cuartos de final',         '+4 pts'),
        ('Bonus Semifinales',              '+5 pts'),
        ('Bonus 3er y 4to puesto',         '+6 pts'),
        ('🏆 Bonus Campeón y subcampeón',  '+7 pts'),
    ]):
        bg = GRIS1 if i%2==0 else BLANCO
        ws.merge_cells(f'A{fila}:G{fila}')
        c = ws.cell(fila, 1, desc)
        c.font=font(size=9); c.fill=fill(bg); c.alignment=alin('left'); c.border=borde()
        ws.merge_cells(f'H{fila}:I{fila}')
        c = ws.cell(fila, 8, pts)
        c.font=font(bold=True,size=9,color=DORADO if pts else NEGRO)
        c.fill=fill(bg); c.alignment=alin(); c.border=borde()
        ws.row_dimensions[fila].height = 17
        fila += 1

    # ── Validación desempate fairplay (1-4) ───────────────────────────────────
    dv_fp = DataValidation(
        type='whole', operator='between', formula1=1, formula2=4,
        allow_blank=True, showInputMessage=True,
        promptTitle='Desempate por Fair Play',
        prompt='Si dos o más equipos quedan empatados en PUNTOS, DG y GF: escribe 1 para el que quieres 1°, 2 para 2°, 3 para 3°, 4 para 4°.',
        showErrorMessage=True,
        errorTitle='Solo 1, 2, 3 o 4',
        error='Escribe 1 (1° lugar), 2 (2° lugar), 3 (3° lugar) o 4 (4° lugar).',
        errorStyle='stop'
    )
    ws.add_data_validation(dv_fp)
    for r in fp_refs:
        dv_fp.add(r)

    # ── Validación numérica en celdas de predicción ───────────────────────────
    print(f'  Aplicando validación en {len(pred_refs)} celdas de predicción...')
    dv = DataValidation(
        type='whole',
        operator='between',
        formula1=0,
        formula2=99,
        allow_blank=True,
        showInputMessage=True,
        promptTitle='Predicción',
        prompt='Escribe el número de goles (0 al 99)',
        showErrorMessage=True,
        errorTitle='Solo números enteros',
        error='Solo puedes escribir números como 0, 1, 2, 3... No letras ni símbolos.',
        errorStyle='stop'
    )
    ws.add_data_validation(dv)
    for r in pred_refs:
        dv.add(r)

    # ── Desbloquear SOLO celdas de marcadores y nombre ───────────────────────
    for r in pred_refs:
        ws[r].protection = Protection(locked=False)

    ws['O1'].protection = Protection(locked=False)
    # pen_refs (penaltis) quedan bloqueadas — todo lo demás también

    # ── Proteger la hoja ──────────────────────────────────────────────────────
    # Sin contraseña: se puede desproteger desde Excel > Revisar > Desproteger hoja
    ws.protection.sheet = True
    ws.protection.selectLockedCells = False    # permitir navegar por toda la hoja
    ws.protection.selectUnlockedCells = False  # pero solo editar las desbloqueadas

    wb.save(ARCHIVO)
    print(f'\n✅ Listo: {ARCHIVO}')
    print('   • Celdas amarillas: predicciones (solo números 0–99)')
    print('   • Celdas lilas (col J): ganador en penaltis en eliminatorias')
    print('   • Celdas moradas (col R): desempate fairplay por grupo (solo 1-4)')
    print('     → Solo llenar si dos o más equipos quedan empatados en pts, DG y GF')
    print('     → Escribe 1 para el equipo que quieres en 1°, 2 para 2°, etc.')
    print('   • Celda dorada O1: tu nombre')
    print('   • Bracket: estructura oficial FIFA 2026 (M73-M104)')
    print('   • Para desproteger: Excel > Revisar > Desproteger hoja (sin contraseña)')

if __name__ == '__main__':
    main()
